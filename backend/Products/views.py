from decimal import Decimal
import re

from django.db import transaction
from rest_framework import generics
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError

from django.db.models import Q

from .models import Product, ProductCategory, UnitOfMeasure
from .serializers import ProductSerializer
from accounts.permissions import IsSuperUserOrReadOnly


def _parse_product_excel(file):
    """
    Parse Excel file (AMARE GOLD LIST style).
    Expected columns: NO.(0), MAKE(1), PRODUCT(2), FMSI(3), APPLICATION(4), POSITION(5), PRICE(6).
    APPLICATION = comma-separated vehicle models the product applies to.
    Yields dicts: {fmsi_number, application, position, unit_price, brand, product_type}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        raise ValidationError("Excel file has no sheets.")

    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_list = list(row or [])
        row_str = [str(c).strip().upper() if c is not None else "" for c in row_list]
        if "FMSI" in row_str and any("APPLICATION" in s for s in row_str) and "PRICE" in row_str:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValidationError(
            "Could not find header row. Expected columns: NO., MAKE, PRODUCT, FMSI, APPLICATION, POSITION, PRICE."
        )

    # Column indices: 0=NO., 1=MAKE, 2=PRODUCT, 3=FMSI, 4=APPLICATION, 5=POSITION, 6=PRICE
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx:
            continue
        row = list(row or [])
        while len(row) < 7:
            row.append(None)

        no_val = row[0]
        make = row[1]  # Brand (TOYOTA, NISSAN, etc.)
        product_type = row[2]  # BRAKEPAD, DISK, etc.
        fmsi = row[3]
        application = row[4]  # Vehicle models, comma-separated
        position = row[5]
        price = row[6]

        # Skip rows without FMSI (brand header rows or empty)
        fmsi_str = (fmsi or "").strip()
        if not fmsi_str:
            continue
        # Skip if no price (incomplete row)
        if price is None or price == "":
            continue

        # Skip if NO. is non-numeric text (brand header row like "TOYOTA" spanning columns)
        try:
            if no_val is not None and str(no_val).strip():
                if not re.match(r"^\d+$", str(no_val).strip()) and not application and not position:
                    continue
        except (TypeError, ValueError):
            pass

        # Normalize application: strip, preserve comma-separated vehicle models
        application_str = (application or "").strip()
        position_str = (position or "").strip()
        brand_str = (make or "").strip() if make else ""
        product_type_str = (product_type or "").strip() if product_type else ""

        # Parse price
        try:
            price_decimal = Decimal(str(price).replace(",", "").strip())
            if price_decimal < 0:
                continue
        except (ValueError, TypeError):
            continue

        yield {
            "fmsi_number": fmsi_str,
            "application": application_str,
            "position": position_str,
            "unit_price": price_decimal,
            "brand": brand_str,
            "product_type": product_type_str,
        }


class ProductImportExcelView(APIView):
    """Import products from Excel file. Expects AMARE-style format."""
    permission_classes = [IsAuthenticated, IsSuperUserOrReadOnly]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get("file") or request.data.get("file")
        if not file:
            raise ValidationError("No file uploaded. Send 'file' in form-data.")
        if not file.name.endswith((".xlsx", ".xls")):
            raise ValidationError("File must be .xlsx or .xls.")

        created = 0
        updated = 0
        skipped = 0
        errors = []

        try:
            rows = list(_parse_product_excel(file))
        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(f"Failed to parse Excel: {e}")

        with transaction.atomic():
            for i, data in enumerate(rows):
                try:
                    fmsi = data["fmsi_number"]
                    application = data["application"]
                    position = data["position"]
                    unit_price = data["unit_price"]
                    brand = data.get("brand", "") or ""
                    product_type = data.get("product_type", "") or ""

                    # Build display name: FMSI - POSITION (or with app snippet)
                    name = f"{fmsi} - {position}" if position else fmsi
                    if application:
                        app_short = application[:80] + "…" if len(application) > 80 else application
                        name = f"{fmsi} - {position} ({app_short})" if position else f"{fmsi} ({app_short})"

                    # Upsert: match by fmsi + application + position
                    existing = Product.objects.filter(
                        fmsi_number=fmsi,
                        application=application,
                        position=position,
                    ).first()

                    sku = fmsi
                    if Product.objects.filter(sku=sku).exclude(pk=existing.pk if existing else -1).exists():
                        sku = f"{fmsi}-{position}-{i}" if position else f"{fmsi}-{i}"

                    if existing:
                        existing.name = name
                        existing.unit_price = unit_price
                        existing.sku = sku
                        existing.brand = brand or existing.brand
                        existing.product_type = product_type or existing.product_type
                        existing.application = application
                        existing.save()
                        updated += 1
                    else:
                        Product.objects.create(
                            name=name,
                            sku=sku,
                            fmsi_number=fmsi,
                            application=application,
                            position=position,
                            unit_price=unit_price,
                            brand=brand,
                            product_type=product_type,
                            category=ProductCategory.SPARE_PART,
                            unit_of_measure=UnitOfMeasure.EACH,
                        )
                        created += 1
                except Exception as e:
                    errors.append({"row": i + 1, "data": data, "error": str(e)})
                    skipped += 1

        return Response({
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:20],  # Limit error details
        })


class ProductSearchView(APIView):
    """GET /api/v1/products/search/?q=... — auto-suggest search by FMSI, position, brand, application.
    Optional ?vehicle=Make Model — filter to products applicable to that vehicle.
    Optional ?site_id=N — restrict to products in that site's inventory (for add-items / parts sale).
    Availability is computed client-side from inventory."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from Inventories.models import Inventory

        q = (request.query_params.get("q") or "").strip()
        vehicle = (request.query_params.get("vehicle") or "").strip()
        site_id = request.query_params.get("site_id")
        limit = min(25, max(5, int(request.query_params.get("limit", 15))))
        qs = Product.objects.filter(is_active=True)

        # Restrict to products in site's inventory when adding items to a sale/service at that site
        if site_id:
            try:
                sid = int(site_id)
                product_ids = Inventory.objects.filter(site_id=sid).values_list("product_id", flat=True)
                qs = qs.filter(id__in=product_ids)
            except (ValueError, TypeError):
                pass

        # Optional: filter by vehicle make/model (application contains it).
        # Only apply when it yields results; otherwise show all site inventory (vehicle filter too strict).
        if vehicle:
            qs_vehicle = qs.filter(application__icontains=vehicle)
            if qs_vehicle.exists():
                qs = qs_vehicle
        if not q:
            # Return initial product list for browse/select (e.g. when adding to inventory).
            # With vehicle/site_id we could add availability; for now return catalog.
            qs = qs.order_by("name")[:limit]
            return Response([{
                "id": p.id,
                "name": p.name,
                "fmsi_number": p.fmsi_number or "",
                "position": p.position or "",
                "application": (p.application or "")[:100],
                "product_type": p.product_type or "",
                "brand": p.brand or "",
                "unit_price": str(p.unit_price),
                "sku": p.sku or "",
                "image_url": request.build_absolute_uri(p.image.url) if p.image else None,
            } for p in qs])
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(fmsi_number__icontains=q)
            | Q(position__icontains=q)
            | Q(brand__icontains=q)
            | Q(application__icontains=q)
            | Q(product_type__icontains=q)
            | Q(sku__icontains=q)
            | Q(part_number__icontains=q)
        ).order_by("name")[:limit]
        return Response([
            {
                "id": p.id,
                "name": p.name,
                "fmsi_number": p.fmsi_number or "",
                "position": p.position or "",
                "application": (p.application or "")[:100],
                "product_type": p.product_type or "",
                "brand": p.brand or "",
                "unit_price": str(p.unit_price),
                "sku": p.sku or "",
                "image_url": request.build_absolute_uri(p.image.url) if p.image else None,
            }
            for p in qs
        ])


class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsSuperUserOrReadOnly]  # Only superusers should access this

    def get_queryset(self):
        qs = super().get_queryset()
        q = (self.request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(name__icontains=q)
                | Q(fmsi_number__icontains=q)
                | Q(position__icontains=q)
                | Q(brand__icontains=q)
                | Q(application__icontains=q)
                | Q(product_type__icontains=q)
                | Q(sku__icontains=q)
                | Q(part_number__icontains=q)
            )
        return qs


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsSuperUserOrReadOnly]
